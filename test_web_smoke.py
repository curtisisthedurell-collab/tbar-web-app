"""
test_web_smoke.py

Smoke/regression tests for the WEB edition's upload plumbing, run headless:

  1. ``parse_cdf_bytes`` produces datasets IDENTICAL to ``parse_cdf_file``
     (same header dict, raw columns, extra columns, timestamps -- everything
     except the recorded source path). Verified against every sample file in
     ``samples/``.
  2. The exporters (``tbar_pdf.build_pdf`` and ``tbar_excel.build_excel``)
     work UNCHANGED when handed in-memory streams (io.BytesIO) -- exactly
     how the Streamlit app uses them -- and produce structurally valid
     output: PDF magic header; xlsx that openpyxl can reopen with the
     expected four sheets and live formulas.

``test_smoke.py`` covers the engine's numeric correctness in depth; this
file covers the web-specific plumbing (byte uploads, in-memory export
streams).

Run:
    python test_web_smoke.py
"""

from __future__ import annotations

import glob
import io
import os

from openpyxl import load_workbook

from tbr_calc import CalibrationSettings, compute_series
from rdf_parser import parse_cdf_bytes, parse_cdf_file
from tbar_excel import (
    CALC_SHEET,
    DERIVED_SHEET,
    FORMULA_SHEET,
    RAW_SHEET,
    build_excel,
)
from tbar_pdf import PlotAxisScale, ReportMetadata, build_pdf

SAMPLES_DIR = os.path.join(os.path.dirname(__file__), "samples")


def approx(a, b, tol=1e-9):
    return abs(a - b) <= tol


def check_dataset_parity(sample_path: str) -> None:
    """parse_cdf_bytes(content) must equal parse_cdf_file(path), except for
    the recorded source path (filename vs full filesystem path)."""
    name = os.path.basename(sample_path)
    with open(sample_path, "rb") as f:
        data = f.read()

    ds_file = parse_cdf_file(sample_path)
    ds_bytes = parse_cdf_bytes(data, source_name=name)

    assert ds_bytes.source_path == name, ds_bytes.source_path
    assert ds_file.header == ds_bytes.header, "header mismatch"
    assert ds_file.is_precalculated == ds_bytes.is_precalculated is True
    assert ds_file.load_raw == ds_bytes.load_raw, "load_raw mismatch"
    assert ds_file.position_raw == ds_bytes.position_raw, "position_raw mismatch"
    assert ds_file.timestamps == ds_bytes.timestamps, "timestamps mismatch"
    assert set(ds_file.extra_columns) == set(ds_bytes.extra_columns), "extra column keys"
    for key in ds_file.extra_columns:
        assert ds_file.extra_columns[key] == ds_bytes.extra_columns[key], f"extra column {key}"
    print(
        f"[OK] {name}: byte-parse identical to path-parse "
        f"({len(ds_bytes.timestamps):,} samples)"
    )


def check_exports_via_streams(sample_path: str) -> None:
    """Both exporters must work unchanged with BytesIO targets and yield
    structurally valid documents."""
    name = os.path.basename(sample_path)
    dataset = parse_cdf_bytes(open(sample_path, "rb").read(), source_name=name)
    calibration = CalibrationSettings.from_dataset(dataset)
    series = compute_series(dataset, calibration)

    # Generic sanity on the computed series.
    n = len(dataset.timestamps)
    assert len(series.depth_m) == n and len(series.qnt_mpa) == n
    assert series.depth_m[0] == 0.0 or approx(series.depth_m[0], 0.0)
    assert all(-1e3 < v < 1e3 for v in series.resistance_mpa), "resistance out of range"

    metadata = ReportMetadata(
        project=dataset.header_get("Project Name", "Project", default=""),
        client="Smoke Test Client",
        location_id=dataset.header_get("Push Name", "Location", default="") or "SMOKE",
        source_filename=name,
        unit_weight_kn_m3=f"{calibration.unit_weight_kn_m3:g}",
        nk_factor=f"{calibration.nk_factor:g}",
        resistance_label="qn,T-bar (MPa)",
    )

    # ---- PDF to an in-memory stream ------------------------------------
    pdf_buf = io.BytesIO()
    build_pdf(
        output_path=pdf_buf,
        metadata=metadata,
        depth_m=series.depth_m,
        resistance_series=series.qnt_mpa,
        elapsed_s=series.elapsed_s,
        depth_scale=PlotAxisScale(),
        time_scale=PlotAxisScale(),
        cycles=series.cycles,
    )
    pdf = pdf_buf.getvalue()
    assert pdf.startswith(b"%PDF"), "output is not a PDF"
    assert len(pdf) > 20_000, f"suspiciously small PDF ({len(pdf)} bytes)"
    print(f"[OK] {name}: build_pdf -> BytesIO produced valid "
          f"{len(pdf):,}-byte PDF")

    # ---- Excel to an in-memory stream -----------------------------------
    xlsx_buf = io.BytesIO()
    build_excel(
        output_path=xlsx_buf,
        dataset=dataset,
        metadata=metadata,
        calibration=calibration,
        series=series,
    )
    workbook = load_workbook(io.BytesIO(xlsx_buf.getvalue()))
    assert workbook.sheetnames == [
        RAW_SHEET, CALC_SHEET, DERIVED_SHEET, FORMULA_SHEET
    ], workbook.sheetnames

    # The Derived Data sheet must carry LIVE formulas (data_type "f"), at
    # least five formula columns per sample row -- proving the auditable
    # live-formula design survived the stream round-trip.
    derived = workbook[DERIVED_SHEET]
    formula_cells = sum(
        1 for row in derived.iter_rows() for cell in row
        if cell.data_type == "f"
    )
    assert formula_cells >= 5 * n, (
        f"expected >= {5 * n} live formulas on '{DERIVED_SHEET}', "
        f"found {formula_cells}"
    )

    # And the Formula Reference sheet must contain NO live formulas (the
    # documented corruption bug the desktop test suite guards against).
    reference = workbook[FORMULA_SHEET]
    for row in reference.iter_rows():
        for cell in row:
            assert cell.data_type != "f", (
                f"{FORMULA_SHEET}!{cell.coordinate} written as a live "
                "formula -- reproduces the Excel repair/corruption bug."
            )
    print(
        f"[OK] {name}: build_excel -> BytesIO produced valid workbook "
        f"(4 sheets, {formula_cells:,} live formulas on '{DERIVED_SHEET}')"
    )


def main():
    samples = sorted(glob.glob(os.path.join(SAMPLES_DIR, "*.cdf")))
    assert samples, f"no .cdf sample files found in {SAMPLES_DIR}"

    for sample in samples:
        check_dataset_parity(sample)
        check_exports_via_streams(sample)

    print("\nALL WEB SMOKE TESTS PASSED")


if __name__ == "__main__":
    main()
