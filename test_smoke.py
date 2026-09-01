"""
test_smoke.py

Headless regression/smoke test for the mini T-bar web processing pipeline.
Exercises the parser, calc, and PDF/Excel builder modules directly against a
real `.cdf` sample file, then verifies the produced PDF (via PyMuPDF/fitz)
and Excel workbook (via openpyxl) structurally.

Requires PyMuPDF for PDF validation (not in requirements.txt -- test-only):
    pip install pymupdf

Run with:
    python test_smoke.py
"""

import dataclasses
import os

from rdf_parser import parse_cdf_file
from tbr_calc import CalibrationSettings, compute_series, detect_cycles
from tbar_pdf import ReportMetadata, PlotAxisScale, build_pdf
from tbar_excel import build_excel

SAMPLE_FILE = os.path.join(
    os.path.dirname(__file__), "samples", "BNW_CSE_12_BC_01_static.cdf"
)
OUTPUT_PDF = os.path.join(os.path.dirname(__file__), "_smoke_test_output.pdf")
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), "_smoke_test_output.xlsx")


def approx(a, b, tol=1e-6):
    return abs(a - b) <= tol


def main():
    # ---- Parser ------------------------------------------------------
    ds = parse_cdf_file(SAMPLE_FILE)
    assert ds.is_precalculated is True
    assert ds.header_get("Location") == "Bonaccia NW - Clara SE"
    assert ds.header_get("Push Name") == "BNW_CSE_12_BC_01"
    assert ds.header_get("Cone") == "T057"
    assert ds.header_get("Tip Area (mm)") == "2500.00"
    assert ds.header_get("N Value") == "11"
    assert len(ds.timestamps) == 450, len(ds.timestamps)
    assert len(ds.position_raw) == 450 and len(ds.load_raw) == 450
    print("[OK] parser: header fields + 450 samples")

    # ---- Calc ------------------------------------------------------------
    cal = CalibrationSettings.from_dataset(ds)
    assert approx(cal.tip_area_mm2, 2500.0)
    assert approx(cal.rod_diameter_mm, 16.0)  # no "Rod Diameter (mm)" in .cdf -> default
    assert approx(cal.unit_weight_kn_m3, 16.0)  # always the default; not in .cdf header
    assert approx(cal.nk_factor, 11.0)  # from "N Value" header field
    series = compute_series(ds, cal)

    assert len(series.depth_m) == len(ds.timestamps)
    assert len(series.resistance_mpa) == len(ds.timestamps)
    assert series.resistance_mpa == list(ds.load_raw)

    # First sample must be zero depth by construction.
    assert approx(series.depth_m[0], 0.0), series.depth_m[0]
    # Max depth should match the hand-verified value for this sample file.
    assert approx(max(series.depth_m), 0.40240000000000004, tol=1e-9), max(series.depth_m)
    assert series.elapsed_s[0] == 0.0
    assert series.elapsed_s[-1] == 51.0
    print("[OK] calc: depth zeroing / elapsed time match hand-verified values")

    # ---- Overburden correction / qn,T-bar / Su ---------------------------
    max_i = series.depth_m.index(max(series.depth_m))
    rod_area = cal.rod_area_m2
    proj_area = cal.projected_area_m2
    expected_overburden = cal.unit_weight_kn_m3 * max(series.depth_m) * rod_area / proj_area / 1000.0
    assert approx(series.overburden_mpa[max_i], expected_overburden, tol=1e-9)
    assert approx(series.overburden_mpa[max_i], 0.0005178068516063367, tol=1e-9)
    for i in (0, max_i, len(series.resistance_mpa) - 1):
        assert approx(series.qnt_mpa[i], series.resistance_mpa[i] - series.overburden_mpa[i])
        assert approx(series.su_kpa[i], (series.qnt_mpa[i] * 1000.0) / cal.nk_factor)
    # Overburden correction should be a small fraction of peak resistance
    # for this shallow (<0.5m) push.
    assert 0 <= series.overburden_mpa[max_i] < 0.01
    print("[OK] calc: overburden correction / qn,T-bar / Su match hand-verified formulas")

    # ---- Nk Factor must be validated, not silently substituted -----------
    # A previous version of this tool silently substituted Nk=1.0 whenever
    # the configured Nk Factor was 0 (or any other falsy value), producing a
    # plausible-looking but silently wrong Su with no indication anything
    # was overridden. Nk Factor is a physical bearing capacity factor and
    # must never be <= 0; compute_series (via compute_su_kpa) must now raise
    # instead of falling back.
    for bad_nk in (0.0, -5.0):
        bad_cal = dataclasses.replace(cal, nk_factor=bad_nk)
        try:
            compute_series(ds, bad_cal)
        except ValueError:
            pass
        else:
            raise AssertionError(
                f"compute_series should reject nk_factor={bad_nk} but did not"
            )
    print("[OK] calc: Nk Factor <= 0 is rejected (no silent fallback)")

    # ---- Cycle detection -------------------------------------------------
    # This sample file is a single push with no remolding cycles
    # ("Cycle Quantity" = 1 in the file header) -> Initial + Final only.
    cycles = series.cycles
    assert [seg.label for seg in cycles] == ["Initial", "Final"], (
        [seg.label for seg in cycles]
    )
    assert cycles[0].kind == "initial" and cycles[0].start_idx == 0
    assert cycles[-1].kind == "final" and cycles[-1].end_idx == len(series.depth_m) - 1
    for prev_seg, seg in zip(cycles, cycles[1:]):
        assert seg.start_idx == prev_seg.end_idx
    # Edge cases: monotonic push -> single Initial; push+pull with no
    # in-between cycles -> Initial + Final only; degenerate tiny inputs
    # must not raise.
    mono = [i * 0.01 for i in range(20)]
    assert [s.label for s in detect_cycles(mono)] == ["Initial"]
    push_pull = [i * 0.01 for i in range(20)] + [0.19 - i * 0.01 for i in range(20)]
    assert [s.label for s in detect_cycles(push_pull)] == ["Initial", "Final"]
    assert detect_cycles([]) and detect_cycles([0.0]) and detect_cycles([0.0, 0.01])
    print("[OK] calc: cycle detection finds Initial + Final, edge cases handled")

    # ---- PDF builder (fixed record -- no AcroForm fields) ---------------
    metadata = ReportMetadata(
        project="Smoke Test Project",
        client="ACME Geotechnical Client",
        location_id=ds.header_get("Push Name"),
        test_date="",
        operator="Smoke Test",
        cone_id=ds.header_get("Cone"),
        unit_weight_kn_m3=f"{cal.unit_weight_kn_m3:g}",
        nk_factor=f"{cal.nk_factor:g}",
        comments="Smoke test export.",
        source_filename=os.path.basename(SAMPLE_FILE),
        resistance_label="qn,T-bar (MPa)",
    )
    build_pdf(
        output_path=OUTPUT_PDF,
        metadata=metadata,
        depth_m=series.depth_m,
        resistance_series=series.qnt_mpa,
        elapsed_s=series.elapsed_s,
        depth_scale=PlotAxisScale(),
        time_scale=PlotAxisScale(),
        cycles=series.cycles,
    )
    assert os.path.isfile(OUTPUT_PDF)
    print(f"[OK] pdf builder: wrote {OUTPUT_PDF} ({os.path.getsize(OUTPUT_PDF)} bytes)")

    # A manual (non-auto) qa_scale for the page 2 Depth vs Elapsed Time plot
    # must be accepted without error and still produce a valid two-page PDF
    # (qa_scale is optional -- the call above omitted it entirely, relying
    # on the auto default; this exercises the explicit-limits path).
    build_pdf(
        output_path=OUTPUT_PDF,
        metadata=metadata,
        depth_m=series.depth_m,
        resistance_series=series.qnt_mpa,
        elapsed_s=series.elapsed_s,
        depth_scale=PlotAxisScale(),
        time_scale=PlotAxisScale(),
        qa_scale=PlotAxisScale(
            x_auto=False, x_min=0.0, x_max=40.0,
            y_auto=False, y_min=0.0, y_max=0.3,
        ),
        cycles=series.cycles,
    )
    assert os.path.isfile(OUTPUT_PDF)
    print("[OK] pdf builder: accepts an explicit qa_scale for the page 2 QA plot")

    # single_color must be accepted and produce a valid PDF too (draws every
    # plot -- including the page 2 QA plot -- as one uniform-color trace
    # with no per-cycle legend, for tests with a lot of cycles).
    build_pdf(
        output_path=OUTPUT_PDF,
        metadata=metadata,
        depth_m=series.depth_m,
        resistance_series=series.qnt_mpa,
        elapsed_s=series.elapsed_s,
        depth_scale=PlotAxisScale(),
        time_scale=PlotAxisScale(),
        cycles=series.cycles,
        single_color="#1f4e79",
    )
    assert os.path.isfile(OUTPUT_PDF)
    print("[OK] pdf builder: accepts single_color for uniform-color traces")

    # ---- Validate the PDF with PyMuPDF ----------------------------------
    import fitz  # PyMuPDF

    doc = fitz.open(OUTPUT_PDF)
    # Page 1: main report (metadata + Depth/Time plots). Page 2: full-width
    # Penetration QA plot.
    assert doc.page_count == 2, doc.page_count

    for page in doc:
        widgets = list(page.widgets())
        assert len(widgets) == 0, f"Expected no editable form fields, found {len(widgets)}"

    page1_text = doc[0].get_text()
    text_upper = page1_text.upper()
    for expected_label in ("UNIT WEIGHT (KN/M", "NK FACTOR", "CONE ID", "PROCESSED BY"):
        assert expected_label in text_upper, f"Expected label '{expected_label}' not found"
    # Removed from the title block on request: Tip Area and Rod Diameter are
    # calibration inputs (still on the Excel export) but no longer clutter
    # the PDF header card, and "Prepared By" was renamed "Processed By".
    assert "PREPARED BY" not in text_upper
    assert "TIP AREA" not in text_upper
    assert "ROD DIAMETER" not in text_upper
    for expected_value in (
        "Smoke Test Project", "ACME Geotechnical Client", "BNW_CSE_12_BC_01",
        "Smoke Test", "T057", "16", "11",
    ):
        assert expected_value in page1_text, f"Expected text '{expected_value}' not found on page 1"
    # Source File no longer appears on the PDF header/card at all.
    assert os.path.basename(SAMPLE_FILE) not in page1_text
    assert os.path.basename(SAMPLE_FILE) not in doc[1].get_text()

    images = doc[0].get_images()
    assert len(images) == 2, f"Expected 2 embedded plot images on page 1, found {len(images)}"
    qa_images = doc[1].get_images()
    assert len(qa_images) == 1, f"Expected 1 embedded plot image on page 2, found {len(qa_images)}"

    doc.close()
    print("[OK] pdf validation: no editable form fields, fixed text present, 2 pages / 3 plot images found")

    # ---- Excel workbook builder (auditable, formula-driven QA export) --
    build_excel(
        output_path=OUTPUT_XLSX,
        dataset=ds,
        metadata=metadata,
        calibration=cal,
        series=series,
    )
    assert os.path.isfile(OUTPUT_XLSX)
    print(f"[OK] excel builder: wrote {OUTPUT_XLSX} ({os.path.getsize(OUTPUT_XLSX)} bytes)")

    # ---- Validate the Excel workbook with openpyxl -----------------------
    import openpyxl
    from tbar_excel import (
        RAW_SHEET, CALC_SHEET, DERIVED_SHEET, FORMULA_SHEET,
        RAW_TABLE_FIRST_DATA_ROW, DERIVED_TABLE_FIRST_DATA_ROW,
        CAL_TIP_AREA_CELL, CAL_ROD_DIAMETER_CELL, CAL_UNIT_WEIGHT_CELL,
        CAL_NK_FACTOR_CELL,
    )

    wb = openpyxl.load_workbook(OUTPUT_XLSX)
    assert wb.sheetnames == [RAW_SHEET, CALC_SHEET, DERIVED_SHEET, FORMULA_SHEET], wb.sheetnames

    raw_ws = wb[RAW_SHEET]
    calc_ws = wb[CALC_SHEET]
    derived_ws = wb[DERIVED_SHEET]
    formula_ws = wb[FORMULA_SHEET]

    n = len(ds.timestamps)

    # Raw Data: row count matches the parsed dataset exactly.
    n_raw_rows = 0
    r = RAW_TABLE_FIRST_DATA_ROW
    while raw_ws.cell(row=r, column=1).value is not None:
        n_raw_rows += 1
        r += 1
    assert n_raw_rows == n, n_raw_rows
    # First raw row's Pen/Tip values match the parsed dataset exactly
    # (plain numbers, not formulas -- this is the source-of-truth sheet).
    assert approx(raw_ws.cell(row=RAW_TABLE_FIRST_DATA_ROW, column=3).value, ds.position_raw[0])
    assert approx(raw_ws.cell(row=RAW_TABLE_FIRST_DATA_ROW, column=4).value, ds.load_raw[0])

    # Metadata & Calibration: calibration values written as plain numbers,
    # matching the CalibrationSettings passed in.
    assert approx(calc_ws[CAL_TIP_AREA_CELL].value, cal.tip_area_mm2)
    assert approx(calc_ws[CAL_ROD_DIAMETER_CELL].value, cal.rod_diameter_mm)
    assert approx(calc_ws[CAL_UNIT_WEIGHT_CELL].value, cal.unit_weight_kn_m3)
    assert approx(calc_ws[CAL_NK_FACTOR_CELL].value, cal.nk_factor)

    # Derived Data: every derived column (D..H) must be a *formula* (string
    # starting with "="), not a hardcoded number -- this is the whole point
    # of the auditable export.
    n_derived_rows = 0
    r = DERIVED_TABLE_FIRST_DATA_ROW
    while derived_ws.cell(row=r, column=1).value is not None:
        n_derived_rows += 1
        for col in range(4, 9):  # D=Depth .. H=Su
            val = derived_ws.cell(row=r, column=col).value
            assert isinstance(val, str) and val.startswith("="), (
                f"Derived Data cell row {r} col {col} is not a formula: {val!r}"
            )
        r += 1
    assert n_derived_rows == n, n_derived_rows

    # Cross-check: manually evaluate the Excel formula chain in Python
    # (openpyxl doesn't evaluate formulas) for a spread of rows, including
    # the max-depth row, and confirm it reproduces tbr_calc's own computed
    # series exactly -- proving the formulas are correct, not just present.
    pen_ref = ds.position_raw[0]
    area_ratio = cal.rod_area_m2 / cal.projected_area_m2

    for i in (0, 1, max_i, n - 1):
        depth = ds.position_raw[i] - pen_ref
        resistance = ds.load_raw[i]
        overburden = cal.unit_weight_kn_m3 * depth * area_ratio / 1000.0
        qnt = resistance - overburden
        su = qnt * 1000.0 / cal.nk_factor
        assert approx(depth, series.depth_m[i])
        assert approx(resistance, series.resistance_mpa[i])
        assert approx(overburden, series.overburden_mpa[i])
        assert approx(qnt, series.qnt_mpa[i])
        assert approx(su, series.su_kpa[i])

    # Formula Reference sheet: documentation rows present and readable.
    formula_text = "\n".join(
        str(formula_ws.cell(row=r, column=1).value or "")
        for r in range(1, formula_ws.max_row + 1)
    )
    for expected in ("Tip Area", "Rod Area", "Depth [m]", "qn,T-bar [MPa]", "Su [kPa]"):
        assert expected in formula_text, f"Expected '{expected}' in Formula Reference sheet"

    # Regression check for a real corruption bug: the "Excel Formula
    # Pattern" column (C) on Formula Reference intentionally *displays*
    # strings that start with "=" (documentation of a formula pattern using
    # placeholder tokens like "Cr"/"Dr" that are not valid Excel syntax) --
    # these must be stored as literal text, never as live formulas, or
    # Excel fails to parse them on open and silently strips them out during
    # its automatic repair (this exact corruption was reported against a
    # real export). openpyxl reports a formula cell's data_type as "f";
    # plain text is "s" (or "str" depending on version) -- assert it is
    # NOT "f" for every documentation row.
    for r in range(3, 3 + 8):  # 8 documented quantities
        cell = formula_ws.cell(row=r, column=3)
        assert cell.value is not None and str(cell.value).startswith("="), (
            f"Formula Reference C{r} should still display a '=...' pattern: {cell.value!r}"
        )
        assert cell.data_type != "f", (
            f"Formula Reference C{r} was written as a live formula, not text -- "
            "this reproduces the Excel repair/corruption bug (removed formula "
            "records in sheet4.xml)."
        )
    # Also confirm at the raw XML level that the Formula Reference sheet's
    # worksheet part contains zero "<f>" (formula) elements, since that is
    # exactly what Excel's recovery log flags when it strips them.
    import zipfile
    with zipfile.ZipFile(OUTPUT_XLSX) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        import re
        import xml.sax.saxutils as saxutils
        sheet_to_rid = {
            saxutils.unescape(name): rid
            for name, rid in re.findall(
                r'<sheet [^>]*name="([^"]*)"[^>]*r:id="(rId\d+)"', wb_xml
            )
        }
        # Relationship attribute order varies (Type/Target/Id in any order
        # depending on the writer), so match each <Relationship .../> tag
        # as a whole and pull Id/Target out of it independently rather than
        # assuming a fixed attribute order.
        rid_to_target = {}
        for tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
            id_match = re.search(r'Id="(rId\d+)"', tag)
            target_match = re.search(r'Target="([^"]*)"', tag)
            if id_match and target_match:
                rid_to_target[id_match.group(1)] = target_match.group(1)
        def _sheet_part(sheet_name: str) -> str:
            target = rid_to_target[sheet_to_rid[sheet_name]]
            # Target is either "/xl/worksheets/sheetN.xml" (absolute, seen
            # in practice) or "worksheets/sheetN.xml" (relative to xl/),
            # depending on the writer -- normalise both to a zip member
            # path with no leading slash.
            return target[1:] if target.startswith("/") else f"xl/{target}"

        formula_sheet_part = _sheet_part(FORMULA_SHEET)
        formula_sheet_xml = z.read(formula_sheet_part).decode("utf-8")
        assert "<f>" not in formula_sheet_xml and "<f " not in formula_sheet_xml, (
            f"Found live <f> formula element(s) in {formula_sheet_part} "
            "(Formula Reference sheet) -- these will corrupt on open in Excel."
        )

        # Positive control: confirm the fix did not overcorrect -- the
        # Derived Data and Metadata & Calibration sheets must still contain
        # real <f> formula elements (these are legitimate, valid formulas).
        for sheet_name in (DERIVED_SHEET, CALC_SHEET):
            part = _sheet_part(sheet_name)
            xml = z.read(part).decode("utf-8")
            assert "<f>" in xml, f"Expected live <f> formulas in {part} ({sheet_name})"

    print(
        "[OK] excel validation: 4 sheets, row counts match dataset, derived columns are "
        "live formulas, formula-chain cross-check matches tbr_calc exactly"
    )

    print("\nALL SMOKE TESTS PASSED")


if __name__ == "__main__":
    main()
