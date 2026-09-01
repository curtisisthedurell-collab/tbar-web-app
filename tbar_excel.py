"""
tbar_excel.py

Builds an auditable Excel workbook (.xlsx) companion export to the PDF lab
report, containing:

  1. "Raw Data"               -- file provenance block followed by the raw
                                 sample table taken directly from the parsed
                                 TbarDataset: Pen (m), Tip (MPa Qc), Sleeve
                                 (MPa), Measured Pore (MPa), TiltX/Y/Combined
                                 Tilt (°). Data is pre-calibrated engineering
                                 units from the .cdf file; no sensor conversion
                                 is applied here.
  2. "Metadata & Calibration" -- ReportMetadata fields and CalibrationSettings
                                 values, plus Derived Constants (Tip Area m²,
                                 Rod Area m², Area Ratio) computed with live
                                 Excel formulas from the calibration cells.
  3. "Derived Data"           -- one row per sample with live Excel formulas
                                 for Depth (m), Resistance (MPa), Overburden
                                 Correction (MPa), qn,T-bar (MPa), Su (kPa),
                                 and Cycle Segment.  Formulas reference the Raw
                                 Data and Metadata & Calibration sheets so a QA
                                 reviewer can edit a calibration value in Excel
                                 and watch every derived row recompute live.
  4. "Formula Reference"      -- human-readable documentation of every formula
                                 used, in both algebraic and literal Excel form.

Depth zeroing uses the first Pen (m) sample as the reference so Excel Depth
matches the Python-computed series exactly (depth_reference_index=0 default).
"""

from __future__ import annotations

from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tbr_calc import CalibrationSettings, ComputedSeries, CycleSegment
from tbr_parser import TbarDataset
from tbar_pdf import ReportMetadata

# ---------------------------------------------------------------------------
# Sheet names
# ---------------------------------------------------------------------------
CALC_SHEET = "Metadata & Calibration"
RAW_SHEET = "Raw Data"
DERIVED_SHEET = "Derived Data"
FORMULA_SHEET = "Formula Reference"

# ---------------------------------------------------------------------------
# Fixed cell addresses on "Metadata & Calibration" sheet.
# The calibration section starts at row 12 (heading) with inputs from B13.
# Derived constants follow from B19.
# Keeping these as module-level constants means the Derived Data formulas and
# the sheet-building code always agree on where each value lives.
# ---------------------------------------------------------------------------
CAL_TIP_AREA_CELL = "B13"       # Tip Area (mm²) — from .cdf "Tip Area (mm)" header
CAL_ROD_DIAMETER_CELL = "B14"   # Rod Diameter (mm)
CAL_UNIT_WEIGHT_CELL = "B15"    # Soil Unit Weight (kN/m³)
CAL_NK_FACTOR_CELL = "B16"      # Nk bearing factor

CAL_TIP_AREA_M2_CELL = "B19"   # Derived: Tip Area (m²) = B13 * 1e-6
CAL_ROD_AREA_CELL = "B20"      # Derived: Rod Area (m²) = PI*(d/2)²*1e-6
CAL_AREA_RATIO_CELL = "B21"    # Derived: Rod Area / Tip Area

# Row numbers for the raw sample data table
RAW_TABLE_HEADER_ROW = 14      # row containing "Row #", "Timestamp", ...
RAW_TABLE_FIRST_DATA_ROW = RAW_TABLE_HEADER_ROW + 1

DERIVED_TABLE_HEADER_ROW = 1
DERIVED_TABLE_FIRST_DATA_ROW = DERIVED_TABLE_HEADER_ROW + 1

# ---------------------------------------------------------------------------
# Shared styling helpers
# ---------------------------------------------------------------------------
HEADING_FILL = PatternFill("solid", fgColor="66FF33")   # house green
HEADING_FONT = Font(color="1A3300", bold=True, size=12)  # dark green text on green
LABEL_FONT = Font(bold=True)
TABLE_HEADER_FILL = PatternFill("solid", fgColor="EFFFEA")  # light green tint
TABLE_HEADER_FONT = Font(bold=True, color="2D7A00")


def _section_heading(ws, cell: str, text: str, span: int = 2) -> None:
    ws[cell] = text
    ws[cell].font = HEADING_FONT
    ws[cell].fill = HEADING_FILL
    row = ws[cell].row
    col = ws[cell].column
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = HEADING_FILL


def _label_value(ws, label_cell: str, label: str, value) -> None:
    ws[label_cell] = label
    ws[label_cell].font = LABEL_FONT
    value_col = ws[label_cell].column + 1
    value_cell = ws.cell(row=ws[label_cell].row, column=value_col)
    value_cell.value = value


def _autosize_columns(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _abs_ref(sheet: str, cell: str) -> str:
    """Absolute sheet-qualified cell reference, e.g. "'Metadata & Calibration'!$B$13"."""
    col = "".join(ch for ch in cell if ch.isalpha())
    row = "".join(ch for ch in cell if ch.isdigit())
    return f"'{sheet}'!${col}${row}"


def _write_formula_text(ws, cell: str, text: str) -> None:
    """Write a string that looks like an Excel formula as plain text.
    Used on the Formula Reference sheet for documentation strings with
    placeholder tokens that are not valid Excel syntax."""
    target = ws[cell]
    target.value = text
    target.data_type = "s"


def _cycle_label_for_row(cycles: List[CycleSegment], row_idx: int) -> str:
    for seg in cycles:
        if seg.start_idx <= row_idx <= seg.end_idx:
            return seg.label
    return ""


def _build_raw_data_sheet(
    wb: Workbook,
    dataset: TbarDataset,
    metadata: ReportMetadata,
    parsed_at: datetime,
) -> None:
    ws = wb.create_sheet(RAW_SHEET)

    # Provenance block
    _section_heading(ws, "A1", "Raw Data Source", span=2)
    _label_value(ws, "A2", "Source File Path", dataset.source_path)
    _label_value(ws, "A3", "Source Filename", metadata.source_filename or "")
    _label_value(ws, "A4", "Parsed At", parsed_at.strftime("%Y-%m-%d %H:%M:%S"))
    _label_value(ws, "A5", "Row Count (samples)", len(dataset.timestamps))
    _label_value(ws, "A6", "Software Version", dataset.header_get("Software Version"))
    _label_value(ws, "A7", "Cone / Test ID",
                 dataset.header_get("Fix Number", "Cone", "Test Number", default=""))
    _label_value(ws, "A8", "Location", dataset.header_get("Location", default=""))
    _label_value(ws, "A9", "Data Type",
                 "Pre-calibrated engineering units from .cdf file")
    _section_heading(ws, "A13", "Pre-calibrated Sample Data", span=9)

    headers = [
        "Row #", "Timestamp", "Pen (m)", "Tip (MPa Qc)",
        "Sleeve (MPa)", "Pore (MPa)", "TiltX (°)", "TiltY (°)", "Combined Tilt (°)",
    ]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=RAW_TABLE_HEADER_ROW, column=col_idx, value=text)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL

    sleeve = dataset.extra_columns.get("sleeve_mpa", [])
    pore = dataset.extra_columns.get("pore_mpa", [])
    tiltx = dataset.extra_columns.get("tiltx_deg", [])
    tilty = dataset.extra_columns.get("tilty_deg", [])
    tilt_c = dataset.extra_columns.get("combined_tilt_deg", [])

    def _get(lst, i):
        return lst[i] if i < len(lst) else None

    n = len(dataset.timestamps)
    for i in range(n):
        r = RAW_TABLE_FIRST_DATA_ROW + i
        ts = dataset.timestamps[i]
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2,
                value=ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "")
        ws.cell(row=r, column=3, value=dataset.position_raw[i])  # Pen (m)
        ws.cell(row=r, column=4, value=dataset.load_raw[i])      # Tip MPa
        ws.cell(row=r, column=5, value=_get(sleeve, i))
        ws.cell(row=r, column=6, value=_get(pore, i))
        ws.cell(row=r, column=7, value=_get(tiltx, i))
        ws.cell(row=r, column=8, value=_get(tilty, i))
        ws.cell(row=r, column=9, value=_get(tilt_c, i))

    ws.freeze_panes = ws.cell(row=RAW_TABLE_FIRST_DATA_ROW, column=1)
    _autosize_columns(ws, {
        "A": 8, "B": 22, "C": 12, "D": 14, "E": 13, "F": 12,
        "G": 10, "H": 10, "I": 16,
    })


def _build_metadata_calibration_sheet(
    wb: Workbook,
    metadata: ReportMetadata,
    calibration: CalibrationSettings,
) -> None:
    ws = wb.create_sheet(CALC_SHEET)

    _section_heading(ws, "A1", "Test Metadata", span=2)
    _label_value(ws, "A2", "Project", metadata.project)
    _label_value(ws, "A3", "Client", metadata.client)
    _label_value(ws, "A4", "Location ID / Box Core No.", metadata.location_id)
    _label_value(ws, "A5", "Test Date", metadata.test_date)
    _label_value(ws, "A6", "Operator", metadata.operator)
    _label_value(ws, "A7", "Source Filename", metadata.source_filename)
    _label_value(ws, "A8", "Resistance Series Displayed", metadata.resistance_label)
    _label_value(ws, "A9", "Comments", metadata.comments)
    assert ws["A9"].row == 9

    _section_heading(ws, "A12", "Calibration / Geometry", span=2)
    assert CAL_TIP_AREA_CELL == "B13"
    _label_value(ws, "A13", "Tip Area (mm²)", calibration.tip_area_mm2)
    _label_value(ws, "A14", "Rod Diameter (mm)", calibration.rod_diameter_mm)
    _label_value(ws, "A15", "Soil Unit Weight (kN/m³)", calibration.unit_weight_kn_m3)
    _label_value(ws, "A16", "Nk Factor (qn,T-bar → Su)", calibration.nk_factor)

    _section_heading(ws, "A18", "Derived Constants (formulas)", span=2)
    assert CAL_TIP_AREA_M2_CELL == "B19"
    ws["A19"] = "Tip Area (m²) = Tip Area (mm²) × 1e-6"
    ws["A19"].font = LABEL_FONT
    ws["B19"] = f"={CAL_TIP_AREA_CELL}*0.000001"

    ws["A20"] = "Rod Area (m²) = PI() × (RodDiameter (mm) / 2)² × 1e-6"
    ws["A20"].font = LABEL_FONT
    ws["B20"] = f"=PI()*({CAL_ROD_DIAMETER_CELL}/2)^2*0.000001"

    ws["A21"] = "Rod Area / Tip Area Ratio"
    ws["A21"].font = LABEL_FONT
    ws["B21"] = f"={CAL_ROD_AREA_CELL}/{CAL_TIP_AREA_M2_CELL}"

    _autosize_columns(ws, {"A": 46, "B": 26})


def _build_derived_data_sheet(
    wb: Workbook,
    dataset: TbarDataset,
    calibration: CalibrationSettings,
    series: ComputedSeries,
) -> None:
    """Live Excel formulas for every derived column, referencing Raw Data
    column C (Pen m) and D (Tip MPa) and the Metadata & Calibration cells.
    A QA reviewer can edit any calibration value and the whole sheet recomputes.

    Depth zeroing uses the first raw Pen (m) sample as the reference, matching
    the Python compute_series default (depth_reference_index=0).
    """
    ws = wb.create_sheet(DERIVED_SHEET)

    headers = [
        "Row #", "Timestamp", "Elapsed (s)",
        "Depth (m)", "Resistance (MPa)", "Overburden (MPa)",
        "qn,T-bar (MPa)", "Su (kPa)", "Cycle Segment",
    ]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=DERIVED_TABLE_HEADER_ROW, column=col_idx, value=text)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL

    n = len(dataset.timestamps)
    # Column C on Raw Data = Pen (m); column D = Tip (MPa Qc)
    pen_first_ref = f"'{RAW_SHEET}'!$C${RAW_TABLE_FIRST_DATA_ROW}"

    unit_weight_ref = _abs_ref(CALC_SHEET, CAL_UNIT_WEIGHT_CELL)
    area_ratio_ref = _abs_ref(CALC_SHEET, CAL_AREA_RATIO_CELL)
    nk_ref = _abs_ref(CALC_SHEET, CAL_NK_FACTOR_CELL)

    for i in range(n):
        r = DERIVED_TABLE_FIRST_DATA_ROW + i
        raw_r = RAW_TABLE_FIRST_DATA_ROW + i
        pen_cell = f"'{RAW_SHEET}'!$C${raw_r}"
        tip_cell = f"'{RAW_SHEET}'!$D${raw_r}"
        ts_cell = f"'{RAW_SHEET}'!$B${raw_r}"

        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"={ts_cell}")

        elapsed = series.elapsed_s[i]
        ws.cell(row=r, column=3, value=elapsed if elapsed is not None else "")

        # D: Depth (m) = Pen[i] - Pen[0]
        depth_cell = f"D{r}"
        ws.cell(row=r, column=4,
                value=f"=({pen_cell}-{pen_first_ref})")

        # E: Resistance (MPa) — reference the Tip MPa column directly
        resistance_cell = f"E{r}"
        ws.cell(row=r, column=5, value=f"={tip_cell}")

        # F: Overburden (MPa) = UnitWeight * Depth * AreaRatio / 1000
        overburden_cell = f"F{r}"
        ws.cell(row=r, column=6,
                value=f"={unit_weight_ref}*{depth_cell}*{area_ratio_ref}/1000")

        # G: qn,T-bar (MPa) = Resistance - Overburden
        qnt_cell = f"G{r}"
        ws.cell(row=r, column=7, value=f"={resistance_cell}-{overburden_cell}")

        # H: Su (kPa) = qn,T-bar * 1000 / Nk
        ws.cell(row=r, column=8, value=f"={qnt_cell}*1000/{nk_ref}")

        ws.cell(row=r, column=9, value=_cycle_label_for_row(series.cycles, i))

    ws.freeze_panes = ws.cell(row=DERIVED_TABLE_FIRST_DATA_ROW, column=1)
    ws.cell(row=DERIVED_TABLE_HEADER_ROW, column=2).number_format = "yyyy-mm-dd hh:mm:ss"
    for r in range(DERIVED_TABLE_FIRST_DATA_ROW, DERIVED_TABLE_FIRST_DATA_ROW + n):
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd hh:mm:ss"
    _autosize_columns(ws, {
        "A": 8, "B": 22, "C": 12, "D": 12, "E": 16,
        "F": 20, "G": 14, "H": 12, "I": 16,
    })


def _build_formula_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(FORMULA_SHEET)

    _section_heading(
        ws, "A1",
        "Formula Reference (see tbr_calc.py for full derivation notes)",
        span=3,
    )

    headers = [
        "Quantity",
        "Algebraic Formula",
        "Excel Formula Pattern (Derived Data sheet, row r)",
    ]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=text)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL

    tip_area_m2_ref = _abs_ref(CALC_SHEET, CAL_TIP_AREA_M2_CELL)
    rod_area_ref = _abs_ref(CALC_SHEET, CAL_ROD_AREA_CELL)
    area_ratio_ref = _abs_ref(CALC_SHEET, CAL_AREA_RATIO_CELL)
    tip_area_ref = _abs_ref(CALC_SHEET, CAL_TIP_AREA_CELL)
    rod_d_ref = _abs_ref(CALC_SHEET, CAL_ROD_DIAMETER_CELL)
    unit_weight_ref = _abs_ref(CALC_SHEET, CAL_UNIT_WEIGHT_CELL)
    nk_ref = _abs_ref(CALC_SHEET, CAL_NK_FACTOR_CELL)
    pen_first = f"'{RAW_SHEET}'!$C${RAW_TABLE_FIRST_DATA_ROW}"

    rows = [
        (
            "Tip Area [m²]",
            "Tip Area [mm²] × 1e-6",
            f"={tip_area_ref}*0.000001",
        ),
        (
            "Rod Area [m²]",
            "pi × (RodDiameter [mm] / 2)² × 1e-6",
            f"=PI()*({rod_d_ref}/2)^2*0.000001",
        ),
        (
            "Rod / Tip Area Ratio",
            "Rod Area [m²] / Tip Area [m²]",
            f"={rod_area_ref}/{tip_area_m2_ref}",
        ),
        (
            "Depth [m]",
            "Pen[i] [m] − Pen[0] [m]",
            f"=('Raw Data'!Cr - {pen_first})",
        ),
        (
            "Resistance [MPa]",
            "Tip (MPa Qc) — taken directly from .cdf data",
            "='Raw Data'!Dr",
        ),
        (
            "Overburden Correction [MPa]",
            "UnitWeight [kN/m³] × Depth [m] × (RodArea / TipArea) / 1000",
            f"={unit_weight_ref}*Dr*{area_ratio_ref}/1000",
        ),
        (
            "qn,T-bar [MPa]",
            "Resistance [MPa] − Overburden Correction [MPa]",
            "=Er - Fr",
        ),
        (
            "Su [kPa]",
            "qn,T-bar [MPa] / Nk × 1000",
            f"=Gr*1000/{nk_ref}",
        ),
    ]

    for r_idx, (name, algebraic, excel) in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=name).font = LABEL_FONT
        ws.cell(row=r_idx, column=2, value=algebraic)
        _write_formula_text(ws, f"C{r_idx}", excel)

    note_row = 3 + len(rows) + 1
    ws.cell(row=note_row, column=1, value=(
        "Note: 'r' denotes the current Derived Data row; 'Cr'/'Dr'/etc. denote "
        "that row's own column C/D/etc on the Derived Data sheet, or the matching "
        "row on the Raw Data sheet where stated. All Derived Data cells are live "
        "Excel formulas — edit any calibration cell on 'Metadata & Calibration' "
        "and the whole sheet recomputes."
    ))
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    ws.row_dimensions[note_row].height = 60

    _autosize_columns(ws, {"A": 30, "B": 50, "C": 60})


def build_excel(
    output_path: str,
    dataset: TbarDataset,
    metadata: ReportMetadata,
    calibration: CalibrationSettings,
    series: ComputedSeries,
) -> None:
    """Build the full auditable Excel workbook and save it to output_path."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _build_raw_data_sheet(wb, dataset, metadata, datetime.now())
    _build_metadata_calibration_sheet(wb, metadata, calibration)
    _build_derived_data_sheet(wb, dataset, calibration, series)
    _build_formula_reference_sheet(wb)

    wb.save(output_path)
